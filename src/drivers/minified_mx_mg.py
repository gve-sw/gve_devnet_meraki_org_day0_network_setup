#!/usr/bin/env python3
"""
Copyright (c) 2024 Cisco and/or its affiliates.
This software is licensed to you under the terms of the Cisco Sample
Code License, Version 1.1 (the "License"). You may obtain a copy of the
License at
https://developer.cisco.com/docs/licenses
All use of the material herein must be in accordance with the terms of
the License. All rights not expressly granted by the License are
reserved. Unless required by applicable law or agreed to separately in
writing, software distributed under the License is distributed on an "AS
IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express
or implied.
"""

__author__ = "Trevor Maco <tmaco@cisco.com>"
__copyright__ = "Copyright (c) 2024 Cisco and/or its affiliates."
__license__ = "Cisco Sample Code License, Version 1.1"

import numpy as np
import openpyxl.worksheet.worksheet

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, borders
from openpyxl.styles.borders import Border
from openpyxl.utils.dataframe import dataframe_to_rows
from rich.console import Console
from rich.panel import Panel

from .driver_interface import ExcelDriverInterface


def process_vlans(vlan_df: pd.DataFrame) -> list[dict]:
    """
    Process each Excel line representing a VLAN, parse it into the correct VLAN JSON structure (including vpn and dhcp tied configurations)
    :param vlan_df: Dataframe representing all lines of vlan config from Excel doc
    :return: List of parsed VLAN dictionaries in the proper format
    """
    parsed_vlans = []

    # Convert each VLAN to a dict and add to the current network
    for _, vlan_row in vlan_df.iterrows():
        vlan_dict = vlan_row.to_dict()

        # Convert Dict to appropriate format (check for minimum fields)
        required_fields = ["name", "subnet", 'applianceIp']
        if not all(field in vlan_dict for field in required_fields):
            continue

        parsed_vlan = {'id': int(vlan_row.iloc[0]), 'name': vlan_dict['name'], 'subnet': vlan_dict['subnet'],
                       'applianceIp': vlan_dict['applianceIp']}

        # Add VPN Section if Applicable
        if 'vpn' in vlan_dict:
            parsed_vlan['_vpn'] = {'useVpn': vlan_dict['vpn']}

        # Add DHCP Section
        if 'dhcpHandling' in vlan_dict:
            parsed_vlan['_dhcp'] = {"dhcpHandling": vlan_dict['dhcpHandling']}

            # Additional DHCP Fields
            if vlan_dict['dhcpHandling'] != 'Do not respond to DHCP requests':
                if 'dnsNameservers' in vlan_dict and vlan_dict['dnsNameservers']:
                    # Split Comma Separated out (if necessary)
                    name_servers = vlan_dict['dnsNameservers'].split(',')

                    if len(name_servers) == 1:
                        parsed_vlan['_dhcp']['dnsNameservers'] = name_servers[0]
                    else:
                        parsed_vlan['_dhcp']['dnsNameservers'] = '\n'.join(name_servers)

                if ('reservedIpRanges - start' in vlan_dict and vlan_dict['reservedIpRanges - start']) and (
                        'reservedIpRanges - end' in vlan_dict and vlan_dict['reservedIpRanges - end']):
                    if "comment" in vlan_dict and vlan_dict['comment']:
                        comment = vlan_dict['comment']
                    else:
                        comment = ""
                    parsed_vlan['_dhcp']['reservedIpRanges'] = [
                        {"start": vlan_dict['reservedIpRanges - start'], "end": vlan_dict['reservedIpRanges - end'],
                         "comment": comment}]

        parsed_vlans.append(parsed_vlan)

    return parsed_vlans


def process_per_port_vlans(per_port_vlan_df: pd.DataFrame) -> list[dict]:
    """
    Process each Excel line representing a Per Port VLAN config, parse it into the correct Per Port VLAN JSON structure
    :param per_port_vlan_df: Dataframe representing all lines of per port vlan config from Excel doc
    :return: List of parsed Per Port VLAN dictionaries in the proper format
    """
    parsed_vlans = []

    # Convert each VLAN to a dict and add to the current network
    for _, vlan_row in per_port_vlan_df.iterrows():
        vlan_dict = vlan_row.to_dict()

        parsed_vlan = {'portId': int(vlan_row.iloc[0]), 'enabled': vlan_dict['enabled'], 'type': vlan_dict['type'],
                       'vlan': vlan_dict['vlan'], 'accessPolicy': vlan_dict['accessPolicy']}

        parsed_vlans.append(parsed_vlan)

    return parsed_vlans


def insert_into_specific_dict_position(current_dict: dict, pos: int, key: str, value: dict | list) -> dict:
    """
    Insert a key, value into a specific position within a dict (ex: claim configuration must be earlier in the dict!)
    :param current_dict: Current dictionary
    :param pos: Position to insert new key value
    :param key: New Entry Key
    :param value: New Entry Value
    :return: New dictionary including newly inserted value
    """
    # List of tuples (key, value)
    items = list(current_dict.items())

    # Insert into specific position
    items.insert(pos, (key, value))

    return dict(items)


def append_df_to_ws_with_headers(worksheet: openpyxl.worksheet.worksheet.Worksheet, dataframe: pd.DataFrame,
                                 start_row: int, include_headers: bool = True, headers: list | None = None,
                                 seperator: bool = False):
    """
    Append DF to OUTPUT sheet, applying headers as well with header formatting
    :param worksheet: Worksheet Object
    :param dataframe: Dataframe with raw data
    :param start_row: Starting row to begin writing at (adds in spacing for readability)
    :param include_headers: Boolean, controls including headers
    :param headers: List of headers to write out to Excel File
    :param seperator: Determine if this is the seperator blank line between runs (special formating)
    """
    # If headers are to be included, add them first
    if include_headers and headers is not None:
        header_row = start_row
        for c_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=c_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
        start_row += 1  # Move the start row for data one row down after headers

    # Blank Line Separator case, apply borders
    if seperator:
        thick_border = borders.Side(style=None, border_style='thin')
        seperator_border = Border(top=thick_border, bottom=thick_border)

    # Append the data
    for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)

            if seperator:
                cell.border = seperator_border


class MinifiedMXMGDriver(ExcelDriverInterface):
    """
    This class creates MX and MG networks with a focus on VLAN creation, WAN Uplink Bandwidth, and DHCP Configurations
    """

    def __init__(self, input_file: str, console: Console):
        """
        Initialize Driver Class
        """
        super().__init__(input_file)

        self.productTypes = ['appliance', 'cellularGateway']
        self.console = console

    def parse_excel_to_json(self) -> dict:
        self.console.print(Panel.fit(f"Parsing Excel File", title="Step 1.5"))

        day0_config = {'networks': []}

        # Determine where the start point is for raw network data ("Networks" row)
        first_col = pd.read_excel(self.input_file, header=None, usecols=[0])
        start_row = 0
        for row in first_col.index:
            if first_col.iloc[row, 0] == 'Networks':
                start_row = row
                break

        # Read the Excel file (only include everything past "Networks")
        df = pd.read_excel(self.input_file, header=None, engine="openpyxl",
                           skiprows=start_row + 1)  # Use header=None since the file does not have a standard header row
        df = df.replace(np.nan, None)

        # Represents current network we are processing from Excel (metadata minimum with productTypes)
        current_network = {'metadata': {"productTypes": self.productTypes}}

        # Structures carried across iterations
        claim_serials = []
        firmware = {"products": {}}
        devices = []
        address = ""

        # Parse row by row (due to complex structure)
        i = 0
        while i < len(df):
            row = df.iloc[i]

            # Check if first column is None (indicates blank row to skip, finalize processing of current network)
            if row[0] is None:
                # Final actions before append...
                if len(claim_serials) > 0:
                    current_network = insert_into_specific_dict_position(current_network, 1, 'claim',
                                                                         {"serials": claim_serials})
                if len(devices) > 0:
                    current_network = insert_into_specific_dict_position(current_network, 2, 'devices', devices)

                current_network['firmware'] = firmware

                # Print all the settings we found...
                self.console.print(
                    f"Found the Following Network Configurations for [blue]{current_network['metadata']['name'] if 'name' in current_network['metadata'] else 'N/A'}[/]: {list(current_network.keys())}")

                # Add to networks list
                day0_config['networks'].append(current_network)

                # Reset Current Network and structures carried across iterations
                current_network = {'metadata': {"productTypes": self.productTypes}}
                claim_serials = []
                firmware = {"products": {}}
                devices = []
                address = ""

                i += 1
                continue

            # Set Column 0 Value to Lower Case (maximize matching chance)
            column_identifier = str(row[0]).lower()
            remaining_row = row[1:]

            # Meta Data Section
            if "name" in column_identifier and remaining_row.first_valid_index():
                current_network['metadata']['name'] = row[remaining_row.first_valid_index()]

            if "timezone" in column_identifier and remaining_row.first_valid_index():
                current_network['metadata']['timeZone'] = row[remaining_row.first_valid_index()]

            # Address Section
            if "address" in column_identifier and remaining_row.first_valid_index():
                address = row[remaining_row.first_valid_index()]

            # Uplink Bandwidth Section
            if "bandwidth" in column_identifier:
                # Add bandwidth structure if not present
                if "traffic_shaping" not in current_network:
                    current_network['traffic_shaping'] = {"_uplink_bandwidth": {"bandwidthLimits": {}}}

                if "1" in column_identifier:
                    # WAN 1
                    current_network['traffic_shaping']["_uplink_bandwidth"]["bandwidthLimits"]['wan1'] = {}

                    if remaining_row.first_valid_index():
                        current_network['traffic_shaping']["_uplink_bandwidth"]["bandwidthLimits"]['wan1']['limitUp'] = \
                            row[remaining_row.first_valid_index()]

                    if remaining_row.last_valid_index():
                        current_network['traffic_shaping']["_uplink_bandwidth"]["bandwidthLimits"]['wan1'][
                            'limitDown'] = row[remaining_row.first_valid_index()]

                elif "2" in column_identifier:
                    # WAN 2
                    current_network['traffic_shaping']["_uplink_bandwidth"]["bandwidthLimits"]['wan2'] = {}

                    if remaining_row.first_valid_index():
                        current_network['traffic_shaping']["_uplink_bandwidth"]["bandwidthLimits"]['wan2'][
                            'limitUp'] = row[remaining_row.first_valid_index()]

                    if remaining_row.last_valid_index():
                        current_network['traffic_shaping']["_uplink_bandwidth"]["bandwidthLimits"]['wan2'][
                            'limitDown'] = row[remaining_row.first_valid_index()]

            # Claim Section (MX and MG)
            if 'serial' in column_identifier:
                # Check side by side columns (do not add firmware value if left blank!)
                serial = row[remaining_row.first_valid_index()]

                if 'firmware' not in serial.lower():
                    claim_serials.append(serial)

                    # Devices Section
                    if address:
                        devices.append({'serial': serial, 'address': address, "tags": ["meraki_script"]})
                    else:
                        devices.append({'serial': serial, "tags": ["meraki_script"]})

                    # Firmware Section
                    firmware_value = row[remaining_row.last_valid_index()]
                    if 'mg' in column_identifier:
                        firmware['products']['cellularGateway'] = {
                            "nextUpgrade": {"toVersion": {"_name_id": firmware_value}}}
                    elif 'mx' in column_identifier:
                        firmware['products']['appliance'] = {"nextUpgrade": {"toVersion": {"_name_id": firmware_value}}}

            # Site to Site VPN
            if 'site-to-site' in column_identifier and remaining_row.first_valid_index():
                # Check if value is hub (full mesh) - only supported at this time
                value = row[remaining_row.first_valid_index()].lower()

                if value == 'hub':
                    current_network['siteToSiteVPN'] = {'mode': value}

            # VLAN Section
            if 'vlan' in column_identifier:
                vlan_headers = row.tolist()
                vlan_data = []

                i += 1  # Move to the next row to start processing VLAN data
                while i < len(df) is not pd.isnull(df.iloc[i][0]) and isinstance(df.iloc[i][0], int):
                    vlan_data.append(df.iloc[i].tolist())
                    i += 1

                # Create VLAN DataFrame and convert to dictionary if vlan_data is not empty
                if vlan_data:
                    vlan_df = pd.DataFrame(vlan_data, columns=vlan_headers)

                    # Process each vlan, convert to appropriate format
                    current_network['vlans'] = process_vlans(vlan_df)

                continue  # Skip the outer loop increment since it's done internally for VLAN rows

            # VLAN Per Port Section
            if 'mx port' in column_identifier:
                vlan_per_port_headers = row.tolist()
                vlan_per_port_data = []

                i += 1  # Move to the next row to start processing VLAN data
                while i < len(df) is not pd.isnull(df.iloc[i][0]) and isinstance(df.iloc[i][0], int):
                    vlan_per_port_data.append(df.iloc[i].tolist())
                    i += 1

                # Create VLAN DataFrame and convert to dictionary if vlan_data is not empty
                if vlan_per_port_data:
                    vlan_per_port_df = pd.DataFrame(vlan_per_port_data, columns=vlan_per_port_headers)

                    # Process each vlan, convert to appropriate format
                    current_network['vlan_per_port'] = process_per_port_vlans(vlan_per_port_df)

                continue  # Skip the outer loop increment since it's done internally for VLAN rows

            i += 1

        # Append Final Network (no reset actions - final addition):

        # Final actions before append...
        if len(claim_serials) > 0:
            current_network = insert_into_specific_dict_position(current_network, 1, 'claim',
                                                                 {"serials": claim_serials})
        if len(devices) > 0:
            current_network = insert_into_specific_dict_position(current_network, 2, 'devices', devices)

        current_network['firmware'] = firmware

        # Print all the settings we found...
        self.console.print(
            f"Found the Following Network Configurations for [blue]{current_network['metadata']['name'] if 'name' in current_network['metadata'] else 'N/A'}[/]: {list(current_network.keys())}")

        # Add to networks list
        day0_config['networks'].append(current_network)

        return day0_config

    def output_results(self, results: list):
        self.console.print(Panel.fit(f"Output Results to Excel", title="Step 3.5"))

        # Open excel, Check if OUTPUT sheet present, otherwise create it
        wb = load_workbook(self.input_file)
        if "OUTPUT" in wb.sheetnames:
            ws = wb["OUTPUT"]
        else:
            ws = wb.create_sheet("OUTPUT")

        # Build DFs (first DF: Network overview, second DF: VLANs)
        for i, network in enumerate(results):
            settings = network['settings']

            # Define and populate network DF
            network_df = {
                "Network Name": [network["_name"]],
                "TimeZone": [""],
                "WAN 1 Bandwidth": [""],
                "WAN 2 Bandwidth": [""],
                "MX Serial": [""],
                "MG Serial": [""],
                "MX Firmware": [""],
                "MG Firmware": [""],
            }

            # Append Timezone
            if 'creation' in settings:
                if settings['creation']['status'] != "Failure":
                    network_df['TimeZone'][0] = settings['creation']['output']['timeZone']
                else:
                    network_df['TimeZone'][0] = "Error (see logs)"

            # Uplink Bandwidth
            if 'traffic_shaping' in settings:
                if settings['traffic_shaping']['status'] != "Failure":
                    if 'uplink_bandwidth' in settings['traffic_shaping']['output']:
                        bandwidth_limits = settings['traffic_shaping']['output']['uplink_bandwidth']['bandwidthLimits']

                        if 'wan1' in bandwidth_limits:
                            network_df['WAN 1 Bandwidth'][
                                0] = f"{bandwidth_limits['wan1']['limitDown']}(down)/{bandwidth_limits['wan1']['limitUp']}(up)"
                        if 'wan2' in bandwidth_limits:
                            network_df['WAN 2 Bandwidth'][
                                0] = f"{bandwidth_limits['wan2']['limitDown']}(down)/{bandwidth_limits['wan2']['limitUp']}(up)"
                else:
                    network_df['WAN 1 Bandwidth'][0] = "Error (see logs)"
                    network_df['WAN 2 Bandwidth'][0] = "Error (see logs)"

            # Firmware Versions
            if 'firmware' in settings:
                if settings['firmware']['status'] != "Failure":
                    for firmware in settings['firmware']['output']:
                        if 'MX' in firmware:
                            network_df['MX Firmware'][0] = firmware
                        if 'MG' in firmware:
                            network_df['MG Firmware'][0] = firmware
                else:
                    network_df['MX Firmware'][0] = "Error (see logs)"
                    network_df['MG Firmware'][0] = "Error (see logs)"

            # Device Serials
            if 'devices' in settings:
                if settings['devices']['status'] != "Failure":
                    for device in settings['devices']['output']:
                        if 'MX' in device['model']:
                            network_df['MX Serial'][0] = device['serial']
                        if 'MG' in device['model']:
                            network_df['MG Serial'][0] = device['serial']
                else:
                    network_df['MX Serial'][0] = "Error (see logs)"
                    network_df['MG Serial'][0] = "Error (see logs)"

            # Define and populate vlan DF
            vlan_df = {
                "VLAN ID": [],
                "Name": [],
                "Subnet": [],
                "Appliance IP": [],
                "DHCP Handling": [],
                "DNS Nameservers": [],
                "Reserved IP Range - Start": [],
                "Reserved IP Range - End": [],
                "Comment": [],
            }

            if 'vlans' in settings:
                if settings['vlans']['status'] != "Failure":
                    vlans = settings['vlans']['output']

                    for vlan in vlans:
                        vlan_df['VLAN ID'].append(vlan['id'])
                        vlan_df['Name'].append(vlan['name'])
                        vlan_df['Subnet'].append(vlan['subnet'])
                        vlan_df['Appliance IP'].append(vlan['applianceIp'])

                        # DHCP
                        vlan_df['DHCP Handling'].append(vlan['dhcpHandling'])
                        vlan_df['DNS Nameservers'].append(vlan['dnsNameservers'].replace("\n", ","))

                        if len(vlan['reservedIpRanges']) > 0:
                            vlan_df['Reserved IP Range - Start'].append(vlan['reservedIpRanges'][0]['start'])
                            vlan_df['Reserved IP Range - End'].append(vlan['reservedIpRanges'][0]['end'])
                            vlan_df['Comment'].append(vlan['reservedIpRanges'][0]['comment'])
                        else:
                            vlan_df['Reserved IP Range - Start'].append("")
                            vlan_df['Reserved IP Range - End'].append("")
                            vlan_df['Comment'].append("")
                else:
                    vlan_df['VLAN ID'] = "Error (see logs)"
                    vlan_df['Name'] = "Error (see logs)"
                    vlan_df['Subnet'] = "Error (see logs)"
                    vlan_df['Appliance IP'] = "Error (see logs)"

                    # DHCP
                    vlan_df['DHCP Handling'] = "Error (see logs)"
                    vlan_df['DNS Nameservers'] = "Error (see logs)"

                    vlan_df['Reserved IP Range - Start'] = "Error (see logs)"
                    vlan_df['Reserved IP Range - End'] = "Error (see logs)"
                    vlan_df['Comment'] = "Error (see logs)"

            # Create the DataFrames
            df1 = pd.DataFrame(network_df)
            df2 = pd.DataFrame(vlan_df)

            # Append an empty row for spacing before (determine if this is the between network seperator (i=0)
            if i == 0:
                append_df_to_ws_with_headers(ws, pd.DataFrame([[" "] * len(df2.columns)]), ws.max_row,
                                             include_headers=False, seperator=True)
            else:
                append_df_to_ws_with_headers(ws, pd.DataFrame([[" "] * len(df2.columns)]), ws.max_row,
                                             include_headers=False)

            # Append network df with its headers
            append_df_to_ws_with_headers(ws, df1, ws.max_row + 1, include_headers=True, headers=list(df1.columns))

            # Append vlan df with its headers
            append_df_to_ws_with_headers(ws, df2, ws.max_row + 1, include_headers=True, headers=list(df2.columns))

        # Save changes
        wb.save(self.input_file)

        self.console.print(f"Successfully wrote results to [yellow]OUTPUT[/] sheet in: {self.input_file}")
